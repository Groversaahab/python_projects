import openpyxl
from openpyxl import Workbook
import os

class question:
    def __init__(self, ques: str, ans:list, genre: str, difficulty: str, neg_mark: bool):
        self.question = ques
        self.answer = ans
        self.genre = genre
        self.difficulty = difficulty
        self.neg_mark = neg_mark
        if difficulty == "hard":
            self.points = 3
        elif difficulty == "medium":
            self.points = 2
        else:
            self.points = 1
        if neg_mark:
            self.neg_score = (self.points)/4
        else:
            self.neg_score = 0

def intro() -> list:
    print("Hi, welcome to quiz maker")
    print("Please fill the asked questions appropriately:")
    print("Name:", end=" ")
    user_name = input()
    print("Genre of the quize you wanna make:(mix for multiple genre)", end=" ")
    quiz_genre = input()
    print("What is the title of your quiz?", end=" ")
    quiz_title = input()
    print("Difficulty of your quiz:(mix for multiple difficulty questions)", end=" ")
    quiz_difficulty = input()
    print("Do you want to specify difficulty rating of each question?(y/n)", end=" ")
    quiz_question_difficulty = input()
    return [user_name, quiz_genre, quiz_title, quiz_difficulty, quiz_question_difficulty]

def excel_file_initialising(user_name: str):
    xlfile = Workbook()
    xlfile.save(filename=f"./quiz_data/{user_name}.xlsx")

def excel_file_format(xlsheet, custom_dif, quiz_genre, quiz_difficulty):
    h1 = ["Quiz Genre", "Quiz difficulty", "Custom question difficulty"]
    if custom_dif:
        h2 = ["Question", "Question difficulty", "Points", "Genre", "Negative marking", "Answer"]
        p1 = [quiz_genre, quiz_difficulty, "Yes"]
    else:
        h2 = ["Question", "Points", "Genre", "Negative marking", "Answer"]
        p1 = [quiz_genre, quiz_difficulty, "No"]
    
    for i in range(len(h1)):
        xlsheet.cell(row=1, column=(i+1)).value = h1[i]
        xlsheet.cell(row=2, column=(i+1)).value = p1[i]
    for i in range(len(h2)):
        xlsheet.cell(row=4, column=(i+1)).value = h2[i]

def question_maker(custom_dif: bool, quiz_difficulty: str, quiz_genre: str) -> list:
    print("Let's start with noting your questions")
    reply = True
    reply_ans = True
    questions = []
    while reply:
        ans = []
        print("Please type in your question:", end=" ")
        ques = input()
        while reply_ans:
            print("Please type in your answer:", end=" ")
            ans.append(input())
            print("Do you want to provide any other answer?(y/n)", end=" ")
            optional_ans = input()
            if optional_ans == "y":
                reply_ans = True
            else:
                reply_ans = False
        reply_ans = True
        if quiz_genre == "mix":
            print("Please type in the question genre:", end=" ")
            ques_genre = input()
        else:
            ques_genre = quiz_genre
        if custom_dif:
            print("Please type in question difficulty:", end=" ")
            ques_dif = input()
        else:
            ques_dif = quiz_difficulty
        print("Do you want to add neg marking to the question?(y/n):", end=" ")
        neg_mark = input()
        if neg_mark == "y":
            neg_mark = True
        else:
            neg_mark = False
        Q = question(ques=ques, ans=ans, genre=ques_genre, difficulty=ques_dif, neg_mark=neg_mark)
        questions.append(Q)
        print("Do you want to add another question?(y/n):", end=" ")
        next_ques = input()
        if next_ques == "y":
            reply = True
        else:
            reply = False
    return questions

def question_saver(questions: list, xlsheet, custom_dif: bool):
    for i in range(len(questions)):
        if custom_dif:
            xlsheet.cell(row=(i+5), column=1).value = questions[i].question
            xlsheet.cell(row=(i+5), column=2).value = questions[i].difficulty
            xlsheet.cell(row=(i+5), column=3).value = questions[i].points
            xlsheet.cell(row=(i+5), column=4).value = questions[i].genre
            xlsheet.cell(row=(i+5), column=5).value = questions[i].neg_score
            for j in range(len(questions[i].answer)):
                xlsheet.cell(row=(i+5), column=j+6).value = questions[i].answer[j]
        else:
            xlsheet.cell(row=(i+5), column=1).value = questions[i].question
            xlsheet.cell(row=(i+5), column=2).value = questions[i].points
            xlsheet.cell(row=(i+5), column=3).value = questions[i].genre
            xlsheet.cell(row=(i+5), column=4).value = questions[i].neg_score
            for j in range(len(questions[i].answer)):
                xlsheet.cell(row=(i+5), column=j+5).value = questions[i].answer[j]

def main():
    [user_name, quiz_genre, quiz_title, quiz_difficulty, custom_question_difficulty] = intro()
    if custom_question_difficulty == "y":
        custom_question_difficulty = True
    else:
        custom_question_difficulty = False

    if os.path.exists(f"./quiz_data/{user_name}.xlsx"):
        xlfile = openpyxl.load_workbook(f"./quiz_data/{user_name}.xlsx")
        sheet = xlfile.create_sheet(title=quiz_title)
    else:
        excel_file_initialising(user_name)
        xlfile = openpyxl.load_workbook(f"./quiz_data/{user_name}.xlsx")
        sheet = xlfile.active
        sheet.title = quiz_title
    
    excel_file_format(sheet, custom_question_difficulty, quiz_genre, quiz_difficulty)

    questions = question_maker(custom_question_difficulty, quiz_difficulty, quiz_genre)

    question_saver(questions, sheet, custom_question_difficulty)

    xlfile.save(filename=f"./quiz_data/{user_name}.xlsx")

if __name__ == "__main__":
    main()